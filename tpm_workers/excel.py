"""
tpm_workers.excel - Reliability Metrics Excel Worker
ref: MASTER_PLAN_v5.md § 11.2.B + Rule #2 (Tool > AI for math)

Pipeline (Researcher -> Analyst -> Coder -> Reviewer):
  1. Researcher: load CM events for equipment
  2. Analyst:    compute MTBF/MTTR/Availability/Pareto/Cost using numpy (NOT AI)
  3. Coder:      build .xlsx with openpyxl (formulas + chart + conditional format)
  4. Reviewer:   re-open .xlsx + sanity-check (formulas eval, no #REF!, no empty refs)
"""
from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from tpm_workers.base import (
    WorkerInput,
    WorkerResult,
    WorkerStep,
    WorkerType,
)
from tpm_workers.data_loader import (
    CMHistoryLoader,
    is_using_dummy_data,
)
from tpm_workers.metrics import (
    cost_summary,
    full_summary,
    pareto_failures,
    severity_breakdown,
)

log = logging.getLogger(__name__)


# ============================================================
# Step 1: Researcher
# ============================================================
def step_researcher(inp: WorkerInput) -> tuple[WorkerStep, dict[str, Any]]:
    step = WorkerStep(name="researcher")
    cm = CMHistoryLoader()
    cm_df = cm.for_equipment(inp.target_subject)

    if not cm_df.empty and "Date" in cm_df.columns and inp.time_range_days:
        from datetime import timedelta
        cutoff = cm_df["Date"].max() - timedelta(days=inp.time_range_days)
        cm_df = cm_df[cm_df["Date"] >= cutoff]

    payload = {
        "equipment_tag": inp.target_subject,
        "events_df": cm_df,  # pandas DataFrame stays in memory
        "n_events": int(len(cm_df)),
        "data_source": str(cm.path) if cm.path else "(none)",
        "is_dummy": is_using_dummy_data(),
    }
    step.output = {
        "n_events": payload["n_events"],
        "data_source": payload["data_source"],
        "is_dummy": payload["is_dummy"],
    }
    step.notes.append(f"loaded {payload['n_events']} CM rows")
    step.finish()
    return step, payload


# ============================================================
# Step 2: Analyst (numpy/pandas - NOT AI)
# ============================================================
def step_analyst(payload: dict[str, Any], obs_days: int) -> tuple[WorkerStep, dict[str, Any]]:
    step = WorkerStep(name="analyst")
    df = payload["events_df"]
    summary = full_summary(df, observation_period_days=obs_days)
    pareto = pareto_failures(df, by="Problem_Reported", top_n=10)

    metrics = {
        **summary,
        "pareto_full": pareto.to_dict("records"),
        "by_severity": severity_breakdown(df),
        "cost": cost_summary(df),
    }
    step.output = {
        "mtbf_hours": metrics["mtbf"].get("mtbf_hours"),
        "mttr_min": metrics["mttr"].get("mttr_min"),
        "availability_pct": metrics["availability_pct"],
        "n_failure_modes": len(pareto),
    }
    step.notes.append(
        f"MTBF={metrics['mtbf'].get('mtbf_hours','n/a')}h, "
        f"MTTR={metrics['mttr'].get('mttr_min','n/a')}min, "
        f"avail={metrics['availability_pct']}%"
    )
    step.finish()
    return step, metrics


# ============================================================
# Step 3: Coder (openpyxl)
# ============================================================
def step_coder(
    payload: dict[str, Any],
    metrics: dict[str, Any],
    output_dir: Path,
    session_id: str,
    obs_days: int,
) -> tuple[WorkerStep, str]:
    step = WorkerStep(name="coder")
    output_dir.mkdir(parents=True, exist_ok=True)
    tag = (payload["equipment_tag"] or "unknown").replace("/", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"reliability_metrics_{tag}_{timestamp}.xlsx"

    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ---- Sheet 1: Summary ----
        ws = wb.active
        ws.title = "Summary"
        header = Font(bold=True, size=14)
        sub_header = Font(bold=True, size=11)
        warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        ws["A1"] = f"Reliability Report: {payload['equipment_tag']}"
        ws["A1"].font = header
        ws.merge_cells("A1:D1")

        ws["A2"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"
        ws["A3"] = f"Session: {session_id}"
        ws["A4"] = f"Source: {payload['data_source']}"
        ws["A5"] = f"Observation window: last {obs_days} days"
        if payload["is_dummy"]:
            ws["A6"] = "WARNING: DUMMY DATA - replace with real data on Day 1 of internship"
            ws["A6"].fill = warn_fill
            ws["A6"].font = Font(bold=True)

        # Metrics block (rows 8-13)
        ws["A8"] = "Key Metrics"
        ws["A8"].font = sub_header
        ws.append([])
        rows = [
            ("Total events", metrics["n_events"]),
            ("MTBF (hours)", metrics["mtbf"].get("mtbf_hours", "n/a")),
            ("MTTR mean (min)", metrics["mttr"].get("mttr_min", "n/a")),
            ("MTTR p90 (min)", metrics["mttr"].get("mttr_p90", "n/a")),
            ("Availability (%)", metrics["availability_pct"]),
            ("Total cost (THB)", metrics["cost"].get("total_cost_thb", 0)),
            ("Avg cost / event (THB)", metrics["cost"].get("avg_cost_per_event_thb", 0)),
        ]
        for label, value in rows:
            ws.append([label, value])

        # ---- Sheet 2: Pareto (top failure modes) ----
        ws2 = wb.create_sheet("Pareto")
        ws2.append(["Failure Mode", "Count", "% of total", "Cumulative %"])
        for cell in ws2[1]:
            cell.font = sub_header
        for row in metrics.get("pareto_full", []):
            ws2.append([
                row.get("Problem_Reported", ""),
                row.get("count", 0),
                row.get("pct", 0),
                row.get("cum_pct", 0),
            ])

        # Excel formula CHECK: total of count column should equal n_events
        last_row = ws2.max_row
        if last_row > 1:
            ws2.cell(row=last_row + 2, column=1, value="Sum check (formula):").font = sub_header
            ws2.cell(row=last_row + 2, column=2, value=f"=SUM(B2:B{last_row})")
            ws2.cell(row=last_row + 3, column=1, value="Expected n_events:").font = sub_header
            ws2.cell(row=last_row + 3, column=2, value=metrics["n_events"])

        # Bar chart
        if last_row > 1:
            chart = BarChart()
            chart.title = "Pareto: Failure Modes"
            chart.x_axis.title = "Failure Mode"
            chart.y_axis.title = "Count"
            data = Reference(ws2, min_col=2, min_row=1, max_row=last_row, max_col=2)
            cats = Reference(ws2, min_col=1, min_row=2, max_row=last_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            ws2.add_chart(chart, "F2")

        # ---- Sheet 3: Raw Events (sample) ----
        ws3 = wb.create_sheet("Raw Events")
        df = payload["events_df"]
        if not df.empty:
            ws3.append(list(df.columns))
            for cell in ws3[1]:
                cell.font = sub_header
            for _, r in df.iterrows():
                ws3.append([
                    r[c].isoformat() if hasattr(r[c], "isoformat") else r[c]
                    for c in df.columns
                ])

        # Auto-width all sheets
        for sh in wb.worksheets:
            for col in sh.columns:
                # Skip merged cells which lack column_letter
                col_letter = None
                max_len = 0
                for cell in col:
                    try:
                        col_letter = cell.column_letter
                    except (AttributeError, ValueError):
                        continue
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                if col_letter:
                    sh.column_dimensions[col_letter].width = min(max_len + 2, 60)

        wb.save(xlsx_path)
        step.output = {"xlsx": str(xlsx_path), "sheets": [s.title for s in wb.worksheets]}
        step.notes.append(f"wrote {xlsx_path.name}")
    except Exception as e:  # noqa: BLE001
        step.finish(success=False, error=f"xlsx build failed: {e}")
        return step, ""
    step.finish()
    return step, str(xlsx_path)


# ============================================================
# Step 4: Reviewer (formula validator)
# ============================================================
def step_reviewer(xlsx_path: str, expected_n: int) -> tuple[WorkerStep, list[str]]:
    step = WorkerStep(name="reviewer")
    findings: list[str] = []
    if not xlsx_path:
        step.finish(success=False, error="no xlsx to review")
        return step, ["no file produced"]

    try:
        from openpyxl import load_workbook

        wb_data = load_workbook(xlsx_path, data_only=False)
        wb_calc = load_workbook(xlsx_path, data_only=True)

        # Check 1: all expected sheets present
        expected_sheets = {"Summary", "Pareto", "Raw Events"}
        actual = set(wb_data.sheetnames)
        missing = expected_sheets - actual
        if missing:
            findings.append(f"missing sheets: {missing}")

        # Check 2: scan for #REF! / #NAME? / #DIV/0! in calc workbook
        for sh_name in wb_calc.sheetnames:
            ws = wb_calc[sh_name]
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str) and value.startswith("#") and value.endswith("!"):
                        findings.append(f"{sh_name}: error cell {value}")
                        break

        # Check 3: row count in Raw Events matches expected
        if "Raw Events" in actual:
            ws_raw = wb_data["Raw Events"]
            actual_rows = ws_raw.max_row - 1  # minus header
            if actual_rows != expected_n:
                findings.append(
                    f"Raw Events row count mismatch: {actual_rows} (expected {expected_n})"
                )

        step.output = {"n_findings": len(findings), "findings": findings}
        step.notes = findings or ["all checks passed"]
        step.finish(success=len([f for f in findings if "error" in f or "missing" in f]) == 0)
    except Exception as e:  # noqa: BLE001
        step.finish(success=False, error=f"reviewer failed: {e}")
        findings.append(str(e))
    return step, findings


# ============================================================
# Top-level
# ============================================================
def run_excel_worker(inp: WorkerInput) -> WorkerResult:
    result = WorkerResult(worker_type=WorkerType.EXCEL)

    # 1. Researcher
    s1, payload = step_researcher(inp)
    result.add_step(s1)
    if payload["n_events"] == 0:
        result.summary = f"No CM history for {inp.target_subject!r} - aborting"
        result.success = False
        return result

    # 2. Analyst (math via numpy)
    s2, metrics = step_analyst(payload, obs_days=inp.time_range_days)
    result.add_step(s2)
    result.metrics = metrics

    # 3. Coder (openpyxl)
    s3, xlsx_path = step_coder(payload, metrics, inp.output_dir, inp.session_id, inp.time_range_days)
    result.add_step(s3)
    if not s3.success:
        result.summary = f"Coder failed: {s3.error}"
        result.success = False
        return result
    result.output_files = [xlsx_path]

    # 4. Reviewer
    s4, findings = step_reviewer(xlsx_path, payload["n_events"])
    result.add_step(s4)
    result.auditor_findings = findings
    result.auditor_passed = s4.success

    result.summary = (
        f"Excel for {inp.target_subject}: "
        f"{payload['n_events']} events, "
        f"MTBF={metrics['mtbf'].get('mtbf_hours','n/a')}h, "
        f"availability={metrics['availability_pct']}%"
    )
    result.confidence = 0.90 if s4.success else 0.60
    result.success = s4.success
    return result
